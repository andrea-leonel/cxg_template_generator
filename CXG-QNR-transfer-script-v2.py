#!/usr/bin/env python
# coding: utf-8
## Instructions

### 1) Extract the html code from the sat page in CXG Live:
> With the S&T pop-up open, right-click 'Inspect' 
> Do ctrl+f  
> Add skipsAndTriggersHistoryContainer in the search bar 
> Right-click the highlighted line
> Copy > copy element > paste it in a notepad
> Save it as html_{protoid}.html - replace protoid with the survey's protoid
> On Jupyter's folder (the same folder where this script is saved), upload the html file. 

### 2) Extract DataList on CXG Connect:
> Go Questionnaire Builder on the left handside menu
> Serch your survey name and select it
> Click Export S&T
> In the excel file downloaded, there will be a tab called 'DataList', right-click it > Move or Copy
> Under To book: click (new book) and save it as Datalist_{protoid} - replace protoid with the survey's protoid
> On Jupyter's folder (the same folder where this script is saved), upload the html file. 

### 3) Add the protoids you'd like to process to the script.
> Below there is a variable called protoid, add the protoids number there, always in this format ['protoid1','protoid2','protoid3',...,'protoid16']
> with the cell below selected, press shift+Enter.
> The templates will be saved in the same folder as the script.

# In[1]:


import pandas as pd
import requests
from bs4 import BeautifulSoup
from IPython.display import display
from io import StringIO
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import traceback


protoids = ['29352','29364','29387','29401','29463','29245','29205','29229','29237','29249','29250','29251','29253','29261','29282']

def generate_sat_template(protoid):
 
    try:
        with open(f"html_{protoid}.html", "r", encoding="utf-8") as file:
            trigger_history_container = file.read()
    except FileNotFoundError:
        print("The file does not exist!")

    soup = BeautifulSoup(trigger_history_container, 'html5lib')

    box_main = soup.find_all('div', {'class': 'box-main'})
    conditions = pd.DataFrame()
    actions = pd.DataFrame()
    group_n = []

    sat_n = 0

    for element in box_main:
        if "(Suspended)" in str(element):
            continue

        sat_n += 1
    
        condition_table = element.find('table', id=lambda x: x and x.startswith('conditionDescriptionTable'))
        block_exclude = element.find_all('div', {'class': 'historySectionLevelsDiv'})
        block_text = []
        tr = condition_table.find_all('tr')

        for row in tr:
        
            condition_text = row.get_text(strip=True)
            exclude_texts = [tag.get_text(strip=True) for tag in block_exclude if tag]

            for text_exclude in exclude_texts:
                if text_exclude in condition_text:
                    condition_text = condition_text.replace(text_exclude, '')

            input_tag = row.find('input', {'name': 'editremove'})
            if input_tag:
                value_attr = input_tag.get('value')
                if value_attr:

                    if '"group":' in value_attr:
                        group_start = value_attr.find('"group":') + len('"group":')
                        group_end = value_attr.find(',', group_start)
                        group_n = value_attr[group_start:group_end].strip()
                    else:
                        group_n = None

            block_text.append({'sat_n': sat_n, 'group_n': group_n, 'condition_text': condition_text})

        df_condition = pd.DataFrame(block_text, columns=block_text[0])
        conditions = pd.concat([conditions,df_condition],ignore_index=True)

        action_table = element.find('table', id=lambda x: x and x.startswith('historyActionTable'))
        if action_table:
            rows = action_table.find_all('tr')
            table_data = []
            for row in rows:
                cols = row.find_all(['th', 'td'])
                cols_text = [col.get_text(strip=True) for col in cols]
                table_data.append(cols_text)

        df_action = pd.DataFrame(table_data[1:], columns=table_data[0])
        df_action['sat_n'] = sat_n
        actions = pd.concat([actions,df_action],ignore_index=True)

    sat = pd.merge(conditions, actions, on='sat_n', how='left')

# Handling group numbering

    sat['group_n'] = sat['group_n'].astype(int) + 1
    sat['group_n'] = sat['group_n'].apply(lambda x: f"Group {int(x)}")

    sat.loc[:,'group_n'] = sat.apply(lambda row: row['condition_text'] if 'Group' in str(row['condition_text']) else row['group_n'],axis=1)
    sat.loc[:,'group_n'] = sat.apply(lambda row: str(row['group_n']).split(' ', 1)[1] if 'AND' in str(row['group_n']) or 'OR' in str(row['group_n']) else row['group_n'],axis=1)

# Tidying up the table

    new_column_names = {
        'sat_n': 'sat_n',
        'condition_text': 'cond_question',
        'Action': 'act_action_question',
        'Question': 'act_parameter',
        'Answers': 'act_action_answer',
        'Comment': 'act_action_comment'
    }
    sat = sat.rename(columns=new_column_names)
    sat = sat[['sat_n','group_n','cond_question', 'act_parameter','act_action_answer', 'act_action_comment','act_action_question']]
    
# Breaking the tables into action and condition to be processed separately

    conditions = sat[['sat_n','group_n','cond_question']]
    actions = sat[['sat_n','act_parameter','act_action_answer', 'act_action_comment','act_action_question']]

    conditions = conditions.drop_duplicates()
    conditions.loc[:,'condition_n'] = conditions.groupby(['sat_n','group_n']).cumcount() + 1
    conditions['condition_n'] = conditions['condition_n'].astype(str)  # Ensure string type
    conditions.loc[:,'condition_n'] = 'Condition ' + conditions['condition_n'].astype(str)

# Condition processing

    conditions = conditions.copy()

    conditions.loc[:,'cond_logical_operator'] = conditions['cond_question'].apply(lambda x: re.search(r'(AND|OR)', str(x)).group(0) if re.search(r'(AND|OR)', str(x)) else None)
    conditions.loc[:,'cond_question_2'] = conditions.apply(lambda row: row['cond_question'].replace(row['cond_logical_operator'], '') if row['cond_logical_operator'] else row['cond_question'],axis=1)
    conditions.loc[:,'cond_operator'] = conditions['cond_question_2'].apply(lambda x: re.search(r'(IS IN|IS NOT IN)', str(x)).group(0) if re.search(r'(IS IN|IS NOT IN)', str(x)) else None)
    conditions.loc[:,'cond_parameter'] = conditions.apply(lambda row: (row['cond_question_2'].split(' [')[0] if isinstance(row['cond_question_2'], str) and ' [' in row['cond_question_2'] else (row['cond_question_2'].split(row['cond_operator'])[0].strip() if row['cond_operator'] and row['cond_operator'] in row['cond_question_2'] else None)),axis=1)
    conditions['cond_parameter_value'] = conditions['cond_question_2'].apply(lambda x: re.search(r'IN\s\((.*?)\)(?=[a-zA-Z]|$)', x).group(1) if isinstance(x, str) and re.search(r'IN\s\((.*?)\)(?=[a-zA-Z]|$)', x) else None)
    conditions.loc[:,'cond_parameter'] = conditions['cond_parameter'].apply(lambda x: x.split(' [')[0] if isinstance(x, str) and ' [' in x else x)

    conditions.loc[:,'cond_parameter_type'] = 'Answer'

    conditions['cond_type'] = 'Condition' 

    conditions.loc[conditions['cond_question'].str.contains('Group', case=False, na=False), 'cond_operator'] = conditions['cond_question'].apply(lambda x: re.search(r'(AND|OR)', str(x)).group(0) if re.search(r'(AND|OR)', str(x)) else "")
    conditions.loc[conditions['cond_question'].str.contains('Group', case=False, na=False),['cond_parameter_type','cond_parameter', 'parameter_value']] = ["","",""]
    conditions.loc[conditions['cond_question'].str.contains('Group', case=False, na=False), 'cond_type'] = "Group"
    conditions.loc[~conditions['cond_question'].str.contains('Group', case=False, na=False), 'cond_parameter_type'] = "Answer"

    conditions['cond_action'] = None
    conditions['cond_action_option'] = None

    conditions = conditions[['sat_n','group_n','condition_n','cond_type','cond_parameter_type','cond_parameter','cond_operator','cond_parameter_value','cond_action','cond_action_option','cond_logical_operator']]

    answer_split = conditions['cond_parameter_value'].str.split(';', expand=True)
    answer_split.columns = [f'cond_parameter_value_{i+1}' for i in range(answer_split.shape[1])]
    conditions = pd.concat([conditions,answer_split], axis=1)

    conditions_unpivot = pd.melt(conditions, id_vars=['sat_n','group_n','condition_n','cond_type','cond_parameter_type','cond_parameter','cond_operator','cond_action','cond_action_option','cond_logical_operator'], value_vars=answer_split.columns, var_name='answer', value_name='parameter_value')

    conditions_final = conditions_unpivot.rename(columns={'parameter_value': 'cond_parameter_value'})
    conditions_final = conditions_final[['sat_n','group_n','condition_n','cond_type','cond_parameter_type','cond_parameter','cond_operator','cond_parameter_value','cond_action','cond_action_option','cond_logical_operator']]
    conditions_final = conditions_final.drop_duplicates()
    conditions_final = conditions_final[~((conditions_final[['cond_parameter_value', 'cond_action', 'cond_action_option']].isna().all(axis=1)) & (conditions_final['cond_type'] != 'Group'))]
    conditions_final = conditions_final.sort_values(by=['sat_n','group_n','condition_n','cond_type'], ascending=[True,True, True, False])

# Action processing

    actions = actions.copy()

    actions.loc[:,'act_parameter'] = actions['act_parameter'].apply(lambda x: x.split(' [')[0])

    actions.loc[:,'Parameter Type_1'] = actions['act_action_answer'].apply(lambda x: 'Answer' if x != "" else None)
    actions.loc[:,'Parameter Type_2'] = actions['act_action_comment'].apply(lambda x: 'Comment' if x != "" else None)
    actions.loc[:,'act_action_comment'] = actions['act_action_comment'].apply(lambda x: 'Change' if x == 'Set Comment(comment: )' else None)

    action_split = actions['act_action_answer'].str.split(r'\)(?=[a-zA-Z]|$)', expand=True)
    action_split.columns = [f'act_action_answer_{i+1}' for i in range(action_split.shape[1])]
    actions = pd.concat([actions, action_split], axis=1)

    action_split2 = pd.DataFrame()
    for i in range(action_split.shape[1]):
        column_name = f'act_action_answer_{i+1}'
        action_split2[[f'act_action_{i+1}', f'act_parameter_value_{i+1}']] = actions[column_name].apply(
            lambda x: pd.Series(str(x).split('(', 1)) if isinstance(x, str) and '(' in str(x) else pd.Series([None, None]))
    actions = actions[['sat_n','act_parameter', 'act_action_comment','act_action_question']]
    actions = pd.concat([actions, action_split2], axis=1)

    unpivoted_1 = pd.melt(actions, id_vars=['sat_n', 'act_parameter'], value_vars=['act_action_comment'], var_name='act_parameter_type', value_name='act_action')
    unpivoted_1 = unpivoted_1[unpivoted_1['act_action'].notna()]
    unpivoted_1.loc[:,"act_parameter_type"] = "Comment"

    unpivoted_2 = pd.melt(actions, id_vars=['sat_n', 'act_parameter'], value_vars=['act_action_question'], var_name='act_parameter_type', value_name='act_action')
    unpivoted_2 = unpivoted_2[unpivoted_2['act_action'].notna()]
    unpivoted_2.loc[:,"act_parameter_type"] = "Question"

    question_answer = pd.concat([unpivoted_1,unpivoted_2], ignore_index=True)
    question_answer.loc[:,"act_parameter_value"] = None
    question_answer = question_answer[['sat_n','act_parameter_type','act_parameter', "act_parameter_value",'act_action']]

    question_answer = question_answer.drop_duplicates()

    unpivoted_dfs = []
    action_columns = [col for col in actions.columns if col.startswith('act_action_') and col.split('_')[-1].isdigit()]
    for act_action in action_columns:
        unpivoted_n = pd.melt(actions, id_vars=['sat_n', 'act_parameter'], value_vars=[act_action], var_name='act_parameter_type', value_name='value')
        unpivoted_dfs.append(unpivoted_n)

    action_unpivoted= pd.concat(unpivoted_dfs, ignore_index=True)
    action_unpivoted = action_unpivoted[action_unpivoted['value'].notna()]
    action_unpivoted.loc[:,'action_n'] = action_unpivoted['act_parameter_type'].apply(lambda x: x.split('_')[-1] if x.split('_')[-1].isdigit() else None)
    action_unpivoted.loc[:,'act_parameter_type'] = action_unpivoted['act_parameter_type'].apply(lambda x: "act_action" if "action" in x else None)
    action_unpivoted.loc[:,'id'] = action_unpivoted['sat_n'].astype(str) + action_unpivoted['act_parameter'].astype(str) + action_unpivoted['action_n'].astype(str)

    unpivoted_dfs = []
    param_columns = [col for col in actions.columns if col.startswith('act_parameter_value_')]
    for act_param in param_columns:
        unpivoted_n = pd.melt(actions, id_vars=['sat_n', 'act_parameter'], value_vars=[act_param], var_name='act_parameter_type', value_name='value')
        unpivoted_dfs.append(unpivoted_n)

    param_unpivoted= pd.concat(unpivoted_dfs, ignore_index=True)
    param_unpivoted = param_unpivoted[param_unpivoted['value'].notna()]
    param_unpivoted.loc[:,'action_n'] = param_unpivoted['act_parameter_type'].apply(lambda x: x.split('_')[-1] if x.split('_')[-1].isdigit() else None)
    param_unpivoted.loc[:,'act_parameter_type'] = param_unpivoted['act_parameter_type'].apply(lambda x: "act_parameter_value" if "parameter_value" in x else None)
    param_unpivoted.loc[:,'id'] = param_unpivoted['sat_n'].astype(str) + param_unpivoted['act_parameter'].astype(str) + param_unpivoted['action_n'].astype(str)

    actions_param = pd.merge(action_unpivoted, param_unpivoted, how='left', on='id',suffixes=('_action', '_parameter'))

    actions_param = actions_param.sort_values(by=['sat_n_action','act_parameter_action','action_n_action'], ascending=[True, True, True])
    actions_param["act_parameter_type"] = "Answer"
    actions_param = actions_param[['sat_n_action','act_parameter_type','act_parameter_action','value_parameter','value_action','action_n_action']]

    actions_param = actions_param.rename(columns={
        'sat_n_action': 'sat_n',
        'act_parameter_type': 'act_parameter_type',
        'act_parameter_action': 'act_parameter',
        'value_parameter': 'act_parameter_value',
        'value_action': 'act_action',
        'action_n_action': 'action_n'
    })

    actions_param = actions_param.drop_duplicates()

    actions_final = pd.concat([question_answer, actions_param], ignore_index=True)

    actions_final.loc[:,"act_type"] = "Action"
    actions_final.loc[:,"act_group_n"] = None
    actions_final.loc[:,"act_condition_n"] = None
    actions_final.loc[:,"act_operator"] = None
    actions_final.loc[:,"act_action_option"] = None
    actions_final.loc[:,"act_logical_operator"] = None

    actions_final = actions_final[['sat_n','act_group_n','act_condition_n','act_type','act_parameter_type','act_parameter','act_operator','act_parameter_value','act_action','act_action_option','act_logical_operator',]]
    actions_final = actions_final.drop_duplicates()

# Merging conditions and actions

    conditions_final = conditions_final.rename(columns={
        'sat_n': 'sat_n',
        'cond_type': 'type',
        'cond_parameter_type': 'parameter_type',
        'cond_parameter': 'parameter',
        'cond_operator': 'operator',
        'cond_parameter_value': 'parameter_value',
        'cond_action': 'action',
        'cond_action_option': 'action_option',
        'cond_logical_operator': 'logical_operator',
    })

    actions_final = actions_final.rename(columns={
        'sat_n': 'sat_n',
        'act_type': 'type',
        'act_parameter_type': 'parameter_type',
        'act_parameter': 'parameter',
        'act_operator': 'operator',
        'act_parameter_value': 'parameter_value',
        'act_action' : 'action',
        'act_action_option': 'action_option',
        'act_logical_operator': 'logical_operator',
    })

    sat_combined = pd.concat([conditions_final, actions_final], ignore_index=True)#
    sat_combined = sat_combined.sort_values(by=['sat_n','group_n','condition_n','type','parameter'], ascending=[True, True, True,False, True])

# Equivalencies of operators and actions

    sat_combined['operator'] = sat_combined['operator'].apply(lambda x: 'in' if x == 'IS IN' else '!in' if x == 'IS NOT IN' else None)

    action_mapping = {
        'show': 'Show',
        'hide': 'Hide',
        'Change': 'Change',
        'Set Answer': 'Set',
        'Clear Answer': 'Clear',
        'Disable Answer': 'Disable',
        'Enable Answer': 'Enable',
        'Show Answer': 'Show',
        'Hide Answer': 'Hide',
    }
    sat_combined['action'] = sat_combined['action'].apply(lambda x: action_mapping.get(x, None))

# Formatting the final table to the template shape

    def add_null_rows(group):
        rule_row = {'sat_n': group['sat_n'].iloc[0], 'type': 'Rule'}
        group_row = {'sat_n': group['sat_n'].iloc[0], 'type': 'Group'}
        return pd.concat([pd.DataFrame([rule_row, group_row]), group])

    sat_final = sat_combined.groupby('sat_n')[sat_combined.columns.tolist()].apply(add_null_rows).reset_index(drop=True)

    sat_final.loc[:,'name'] = sat_final.apply(lambda row: sat_final.loc[row.name + 1, 'group_n'] if row['type'] == 'Group' and row.name + 1 < len(sat_final) else (row['condition_n'] if row['type'] == 'Condition' else None), axis=1)

    sat_final = sat_final[['type','name','parameter_type','parameter','operator','parameter_value','action','action_option','logical_operator']]

# Looking up question codes and replacing

    datalist = pd.read_excel(f'datalist_{protoid}.xlsx', sheet_name='DataLists')

    sat_final['parameter'] = sat_final['parameter'].apply(
        lambda param: next(
            (row for row in datalist.iloc[:, 0] if str(param) in str(row)),
            param 
        )
    )

    sat_final = sat_final.rename(columns={
            'name': 'Name',
            'parameter_type': 'Parameter Type',
            'parameter': 'Parameter',
            'operator': 'Statement Operator',
            'parameter_value': 'Parameter Value',
            'action': 'Action',
            'action_option': 'Action Option',
            'logical_operator': 'Logical Operator'
        })

# Processing the Excel file

    wb = Workbook()
    ws = wb.active

    for col_idx, column in enumerate(sat_final.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column) 
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sat_final = sat_final.reset_index(drop=True)

    for row_idx, row in sat_final.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    colors = {
        'Rule': 'D88A55',
        'Group': '00B09B',
        'Condition': '9DD3C9',
        'Action': 'D9D9D9'
    }

    for row_idx, row in sat_final.iterrows():
        fill_color = colors.get(row['type'], None)
        if fill_color:
            for col_idx in range(1, len(sat_final.columns) + 1):
                ws.cell(row=row_idx + 2, column=col_idx).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )

    for column_cells in ws.columns:
        ws.column_dimensions[column_cells[0].column_letter].width = 20

    wb.save(f'template_{protoid}.xlsx')

for protoid in protoids:
    try:
        generate_sat_template(protoid)
        print(f"Processed and saved template files for: {protoid}")
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        line_number = tb[-1].lineno 
        print(f"Error processing {protoid}: {e}")


# In[ ]:





# In[ ]:




